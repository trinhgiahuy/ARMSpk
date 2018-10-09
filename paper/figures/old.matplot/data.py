import openpyxl
import urllib
import tempfile
import os

#-- Settings --#

XLSX_LINK = "https://docs.google.com/spreadsheets/d/1un0TIi31LXI9yURmwobPkCOatNXTvVPofXbEu_W-SvA/export?format=xlsx&id=1un0TIi31LXI9yURmwobPkCOatNXTvVPofXbEu_W-SvA"
PERSONS_TO_PLOT = [ "Haoyu", "Tsuji", "Tsuchikawa", "Yashima", "Matsumura" ]
PERSONS = PERSONS_TO_PLOT + [ "Jens" ]

MAX_COL=100
MAX_ROW=50

TITLE_XEON='Final_Xeon'
TITLE_KNL='Final_KNL_cache mode'
TITLE_KNM='Final_KNM_cache mode'

# Result of `./omp-stream-arch --triad-only --float -s $((1000*1000*200))`
# (BabelStream-3.3 compiled by Intel Compiler)
PEAKBW_XEON=46.259
PEAKBW_KNL=288.757 # on HIGH frequency node
PEAKBW_KNM=264.844 # on HIGH frequency node

# Result of `cpuinfo g | grep -v package | grep Cores`
CORE_XEON=12
CORE_KNL=64
CORE_KNM=72

BOTTOM=0

def try_float(value):
    try:
        return float(value)
    except ValueError:
        return BOTTOM

COL_APP=0
COL_TITLE=MAX_COL-1
COL_PERSON=1

def get_flops(row):
    val = search_col(row, "Measured Performance (GFLOPS/s) SP/DP")
    if val is None:
        return BOTTOM
    return try_float(val)

def get_col(row, label):
    val = search_col(row, label)
    if val is None:
        return BOTTOM
    return try_float(val)

def get_bandwidth(row):
    return get_col(row, "Measured Memory Throughput SYSTEM (GB/Sec)")

def get_hitrate(row):
    return get_col(row, "Hit rate (%) of LLC (Last Level Cache)")

def get_core_used(row):
    return get_col(row, "Cores used (min(#MPI*#OMP,#CORE)")

def get_solver_rate(row):
    return get_col(row, "Part of runtime spent in solver (in percentage)")

def get_energy(row):
    return get_col(row, "Power (W)")

def get_bf(row):
    b = get_bandwidth(row)
    f = get_flops(row)
    if b > 0 and f > 0:
        return b / f
    return BOTTOM

def get_ratio(row):
    val_fp64 = search_col(row, "% of FP64 instructions")
    val_fp32 = search_col(row, "% of FP32 instructions")
    val_int  = search_col(row, "% of INT instructions")
    if val_fp64 is None or val_fp32 is None or val_int is None:
        return (BOTTOM, BOTTOM, BOTTOM)
    return (try_float(val_fp64),
            try_float(val_fp32),
            try_float(val_int))

#--------------#

def temp_dl(dl_link, ext):
    with tempfile.NamedTemporaryFile(delete=False) as temp:
        temp.write(urllib.request.urlopen(dl_link).read())
        loc = temp.name

    new_loc = loc + "." + ext
    os.rename(loc, new_loc)
    return new_loc

XLSX_FILE = temp_dl(XLSX_LINK, "xlsx")
WB = openpyxl.load_workbook(XLSX_FILE, data_only=True)

# RESULT_DIC[NAME OF APP][XEON OR KNL OR KNM][COLUMNS]
RESULT_DIC = {}

APP_TO_LABEL = {
    'AMG': 'AMG',
    'CANDLE Benchmarks': 'CAND',
    'CoMD': 'CoMD',
    'Laghos': 'Lagh',
    'MACSio': 'MACS',
    'miniAMR': 'mAMR',
    'miniFE': 'mFE',
    'miniTri': 'mTRI',
    'Nekbone': 'Nekb',
    'SW4lite': 'SW4l',
    'SWFFT': 'SWFF',
    'XSBench': 'XSBe',
    'CCS QCD': 'QCD',
    'FFVC-MINI': 'FFVC',
    'NICAM-DC-MINI': 'NICA',
    'mVMC-MINI': 'mVMC',
    'NGS Analyzer-MINI': 'NGSA',
    'MODYLAS-MINI': 'MODY',
    'NTChem-MINI': 'NTCh',
    'FFB-MINI': 'FFB',
    'HPL': 'HPL',
    'HPCG': 'HPCG',
    'BabelStream 2GB': 'BabelStream 2GB',
    'BabelStream 14GB': 'BabelStream 14GB'
}

for ws in WB:
    for row in ws.iter_rows(min_row=1, max_col=MAX_COL, max_row=MAX_ROW):
        if row[COL_PERSON].value in PERSONS:
            app = APP_TO_LABEL[row[COL_APP].value]
            if ws.title == WB.sheetnames[0]:
                RESULT_DIC[app] = {}
            RESULT_DIC[app][ws.title] = row
            RESULT_DIC[app][ws.title][COL_TITLE].value = ws.title # Only for search_col

def delete_whitespace(string):
    return ''.join(string.split())

# Without whitespaces
def startswith_wo_ws(a, b):
    return delete_whitespace(a).startswith(delete_whitespace(b))

# Find the position of the columns using the first row of the sheet
def search_col(row, label):
    title = row[COL_TITLE].value
    n = -1
    for ws in WB:
        if ws.title == title:
            for idx, cell in enumerate(next(ws.rows)):
                if not isinstance(cell, type(None)) and startswith_wo_ws(cell.value, label):
                    n = idx
                    break
            if n != -1:
                return row[n].value

    print('NOT FOUND: ' + label + ' of ' + title)
    return None

def is_plotted(rows):
    return rows[TITLE_XEON][COL_PERSON].value in PERSONS_TO_PLOT
