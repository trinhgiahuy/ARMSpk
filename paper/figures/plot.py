#!/usr/bin/env python3

import matplotlib
matplotlib.use("Agg")

import warnings
warnings.filterwarnings("ignore",category=matplotlib.cbook.mplDeprecation)

import openpyxl
import urllib
import tempfile
import os
import re
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import seaborn as sns
import numpy as np

#-- Settings --#

XLSX_LINK = "https://docs.google.com/spreadsheets/d/1un0TIi31LXI9yURmwobPkCOatNXTvVPofXbEu_W-SvA/export?format=xlsx&id=1un0TIi31LXI9yURmwobPkCOatNXTvVPofXbEu_W-SvA"
PERSONS = [ "Haoyu", "Tsuji", "Tsuchikawa", "Yashima", "Matsumura", "Jens" ]

MAX_COL=50
MAX_ROW=50

COL_APP=0
COL_PERSON=1
COL_FLOPS=4
COL_TFLOP=4
COL_BANDWIDTH=8
COL_TOTTIME=13
COL_SOLTIME=14
COL_HITRATE=18
COL_BOUND=24
COL_FP64=28
COL_FP32=30
COL_INT=32
COL_COREUSED=37
COL_ENERGY=36

TITLE_XEON='Final_Xeon'
TITLE_KNL='Final_KNL_cache mode'
TITLE_KNM='Final_KNM_cache mode'

# Result of `mpiexec -n 1 -genv KMP_AFFINITY=scatter -genv OMP_NUM_THREADS=<CORES> /usr/bin/numactl --membind=1 ./BabelStream/omp-stream -s $((1000*1000*600)) -n 10 --triad-only --float`
# (BabelStream-3.3 compiled by Intel Compiler)
PEAKBW_XEON=121.510
PEAKBW_KNL=438.809 # on HIGH frequency node
PEAKBW_KNM=429.837 # on HIGH frequency node

# Result of `cpuinfo g | grep -v package | grep Cores`
CORE_XEON=24
CORE_KNL=64
CORE_KNM=72

BOTTOM=np.nan #-5

def try_float(value):
    try:
        return float(value)
    except ValueError:
        return BOTTOM

def get_flops(row):
    val = row[COL_FLOPS].value
    if val is None:
        return BOTTOM
    if isinstance(val, float):
        return val
    return try_float(re.sub(r'\(.*\)', '', val))

def get_tflop(row):
    val = row[COL_TFLOP].value
    if val is None:
        return BOTTOM
    if isinstance(val, float):
        return val
    return try_float(re.sub(r'\(.*\)', '', val))

def get_col(row, col):
    val = row[col].value
    if val is None:
        return BOTTOM
    return try_float(val)

def get_bandwidth(row):
    return get_col(row, COL_BANDWIDTH)

def get_boundedness(row):
    return get_col(row, COL_BOUND)

def get_hitrate(row):
    return get_col(row, COL_HITRATE)

def get_core_used(row):
    return get_col(row, COL_COREUSED)

def get_total_time(row):
    return get_col(row, COL_TOTTIME)

def get_solver_time(row):
    return get_col(row, COL_SOLTIME)

def get_energy(row):
    return get_col(row, COL_ENERGY)

def get_bf(row):
    b = get_bandwidth(row)
    f = get_flops(row)
    if b > 0 and f > 0:
        return b / f
    return BOTTOM

def get_ratio(row):
    val_fp64 = row[COL_FP64].value
    val_fp32 = row[COL_FP32].value
    val_int  = row[COL_INT].value
    if val_fp64 is None or val_fp32 is None or val_int is None:
        return (BOTTOM, BOTTOM, BOTTOM)
    return (try_float(val_fp64),
            try_float(val_fp32),
            try_float(val_int))

#--------------#

def temp_dl(dl_link, ext):
    with tempfile.NamedTemporaryFile(delete=False) as temp:
        temp.write(urllib.request.urlopen(dl_link).read())
        loc = temp.name

    new_loc = loc + "." + ext
    os.rename(loc, new_loc)
    return new_loc

XLSX_FILE = temp_dl(XLSX_LINK, "xlsx")
wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)

result_dic = {}

for ws in wb:
    for row in ws.iter_rows(min_row=1, max_col=MAX_COL, max_row=MAX_ROW):
        if row[COL_PERSON].value in PERSONS:
            if ws.title == wb.sheetnames[0]:
                result_dic[row[COL_APP].value] = {}
            result_dic[row[COL_APP].value][ws.title] = row

def figure_setup(title, ylabel, figsize, grid_interval):
    plt.figure(figsize=figsize)
    sns.set('paper', 'whitegrid', 'deep',
            font_scale=1.5,
            rc={"lines.linewidth": 2,
                'grid.linestyle': '--'})
    plt.title(title, fontsize=20, fontweight='bold')
    plt.ylabel(ylabel)
    #plt.axes().yaxis.set_major_locator(ticker.MultipleLocator(grid_interval))
    plt.axes().yaxis.set_major_locator(ticker.MaxNLocator(12))

def figure_save(filename):
    plt.ylim(ymin=0)
    plt.xticks(rotation=60)
    plt.legend(loc='center left',
               bbox_to_anchor=(1, 0.5))
    plt.tight_layout()
    plt.savefig(filename)
    plt.close()

figure_setup("TEMP", "", (10,10), 1)
colors=plt.rcParams['axes.prop_cycle'].by_key()['color']
COLOR_XEON=colors[2]
COLOR_KNL=colors[0]
COLOR_KNM=colors[1]

# Differences in performance (%)
figure_setup("Time-to-solution [KNL=blue; KNM=green] relative to dual-socket Intel Broadwell-EP", "Speedup", (18,6), 10)
colors=plt.rcParams['axes.prop_cycle'].by_key()['color']

perf_app = []
perf_knl = []
perf_knm = []
for app, rows in result_dic.items():
    xeon_flops = get_solver_time(rows[TITLE_XEON]) #get_flops(rows[TITLE_XEON])
    knl_flops = get_solver_time(rows[TITLE_KNL]) #get_flops(rows[TITLE_KNL])
    knm_flops = get_solver_time(rows[TITLE_KNM]) #get_flops(rows[TITLE_KNM])

    perf_app.append(app[0:7])
    perf_knl.append(xeon_flops / knl_flops) #(knl_flops / xeon_flops * 100) if (xeon_flops > 0 and knl_flops > 0) else BOTTOM)
    perf_knm.append(xeon_flops / knm_flops) #(knm_flops / xeon_flops * 100) if (xeon_flops > 0 and knm_flops > 0) else BOTTOM)

#plt.ylim(ymax=1000)
plt.axhline(y=1, color=COLOR_XEON, ls='--') #y=100, color=COLOR_XEON, ls='--')
perf_app0=perf_app[:]
plt.xticks(list(range(len(perf_app))), perf_app)
perf_app=list(range(len(perf_app)))
plt.plot(perf_app, perf_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(perf_app, perf_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("performance.png")

figure_setup("Performance compared to Xeon", "(%)", (15,6), 10)
colors=plt.rcParams['axes.prop_cycle'].by_key()['color']
plt.axhline(y=100, color=COLOR_XEON, ls='--')
plt.xticks(list(range(len(perf_app0))), perf_app0)
perf_app=list(range(len(perf_app0)))
plt.ylim(ymax=500)
plt.plot(perf_app, perf_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(perf_app, perf_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)
figure_save("performance_zoom.png")


# Differences in performance per core (%)
figure_setup("Performance (per core) compared to Xeon", "(%)", (15,6), 10)

perf_app = []
perf_knl = []
perf_knm = []
for app, rows in result_dic.items():
    xeon_flops = get_flops(rows[TITLE_XEON]) / CORE_XEON
    knl_flops = get_flops(rows[TITLE_KNL]) / CORE_KNL
    knm_flops = get_flops(rows[TITLE_KNM]) / CORE_KNM

    perf_app.append(app[0:5])
    perf_knl.append((knl_flops / xeon_flops * 100) if (xeon_flops > 0 and knl_flops > 0) else BOTTOM)
    perf_knm.append((knm_flops / xeon_flops * 100) if (xeon_flops > 0 and knm_flops > 0) else BOTTOM)

plt.axhline(y=100, color=COLOR_XEON, ls='--')
plt.xticks(list(range(len(perf_app))), perf_app)
perf_app=list(range(len(perf_app)))
plt.plot(perf_app, perf_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(perf_app, perf_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("performance_per_core.png")


# Differences in performance per core used (%)
figure_setup("Performance (per core used) compared to Xeon", "(%)", (15,6), 10)

perf_app = []
perf_knl = []
perf_knm = []
for app, rows in result_dic.items():
    xeon_core_used = get_core_used(rows[TITLE_XEON])
    knl_core_used = get_core_used(rows[TITLE_KNL])
    knm_core_used = get_core_used(rows[TITLE_KNM])
    xeon_flops = get_flops(rows[TITLE_XEON]) / xeon_core_used
    knl_flops = get_flops(rows[TITLE_KNL]) / knl_core_used
    knm_flops = get_flops(rows[TITLE_KNM]) / knm_core_used

    perf_app.append(app[0:5])
    perf_knl.append((knl_flops / xeon_flops * 100) if (xeon_core_used > 0 and knl_core_used > 0 and xeon_flops > 0 and knl_flops > 0) else BOTTOM)
    perf_knm.append((knm_flops / xeon_flops * 100) if (xeon_core_used > 0 and knm_core_used > 0 and xeon_flops > 0 and knm_flops > 0) else BOTTOM)

plt.axhline(y=100, color=COLOR_XEON, ls='--')
plt.xticks(list(range(len(perf_app))), perf_app)
perf_app=list(range(len(perf_app)))
plt.plot(perf_app, perf_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(perf_app, perf_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("performance_per_core_used.png")


# FLOPS
figure_setup("Floating Point Op/s (SP+DP)", "Gflop/s", (10,10), 10)

flops_app = []
flops_xeon = []
flops_knl = []
flops_knm = []
for app, rows in result_dic.items():
    flops_app.append(app[0:5])
    flops_xeon.append(get_flops(rows[TITLE_XEON]))
    flops_knl.append(get_flops(rows[TITLE_KNL]))
    flops_knm.append(get_flops(rows[TITLE_KNM]))

flops_app0=flops_app[:]
plt.xticks(list(range(len(flops_app))), flops_app)
flops_app=list(range(len(flops_app)))
plt.plot(flops_app, flops_xeon, ".", color=COLOR_XEON, label="XEON", markersize=30)
plt.plot(flops_app, flops_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(flops_app, flops_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("flops.png")

figure_setup("Floating Point Op/s (SP+DP)", "Gflop/s", (10,10), 10)
plt.xticks(list(range(len(flops_app0))), flops_app0)
flops_app=list(range(len(flops_app0)))
plt.ylim(ymax=300)
plt.plot(flops_app, flops_xeon, ".", color=COLOR_XEON, label="XEON", markersize=30)
plt.plot(flops_app, flops_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(flops_app, flops_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)
figure_save("flops_zoom.png")


# FLOPS per core
figure_setup("Floating Point Op/s (SP+DP) per core", "Gflops/s", (10,10), 1)

flops_app = []
flops_xeon = []
flops_knl = []
flops_knm = []
for app, rows in result_dic.items():
    flops_app.append(app[0:5])
    flops_xeon.append(get_flops(rows[TITLE_XEON]) / CORE_XEON)
    flops_knl.append(get_flops(rows[TITLE_KNL]) / CORE_KNL)
    flops_knm.append(get_flops(rows[TITLE_KNM]) / CORE_KNM)

#plt.plot(flops_app, flops_xeon, ".", color=COLOR_XEON, label="XEON", markersize=30)
plt.xticks(list(range(len(flops_app))), flops_app)
flops_app=list(range(len(flops_app)))
plt.plot(flops_app, flops_xeon, ".", color=COLOR_XEON, label="XEON", markersize=30)
plt.plot(flops_app, flops_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)
plt.plot(flops_app, flops_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("flops_per_core.png")


# Bandwidth
figure_setup("Memory Throughput", "(GB/sec)", (10,10), 10)

bw_app = []
bw_xeon = []
bw_knl = []
bw_knm = []
for app, rows in result_dic.items():
    bw_app.append(app[0:5])
    bw_xeon.append(get_bandwidth(rows[TITLE_XEON]))
    bw_knl.append(get_bandwidth(rows[TITLE_KNL]))
    bw_knm.append(get_bandwidth(rows[TITLE_KNM]))

plt.xticks(list(range(len(bw_app))), bw_app)
bw_app=list(range(len(bw_app)))

plt.axhline(y=PEAKBW_XEON, color=COLOR_XEON, ls='--')
plt.plot(bw_app, bw_xeon, ".", color=COLOR_XEON, label="Xeon", markersize=30)

plt.axhline(y=PEAKBW_KNL, color=COLOR_KNL, ls='--')
plt.plot(bw_app, bw_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)

plt.axhline(y=PEAKBW_KNM, color=COLOR_KNM, ls='--')
plt.plot(bw_app, bw_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("bandwidth.png")


if 0:
    # Bandwidth per core
    figure_setup("Memory Throughput per Core", "(GB/sec)", (10,10), 0.25)

    bw_app = []
    bw_xeon = []
    bw_knl = []
    bw_knm = []
    for app, rows in result_dic.items():
        bw_app.append(app[0:5])
        bw_xeon.append(get_bandwidth(rows[TITLE_XEON]) / CORE_XEON)
        bw_knl.append(get_bandwidth(rows[TITLE_KNL])  / CORE_KNL)
        bw_knm.append(get_bandwidth(rows[TITLE_KNM])  / CORE_KNM)

    plt.xticks(list(range(len(bw_app))), bw_app)
    bw_app=list(range(len(bw_app)))

    plt.axhline(y=PEAKBW_XEON / CORE_XEON, color=COLOR_XEON, ls='--')
    plt.plot(bw_app, bw_xeon, ".", color=COLOR_XEON, label="Xeon", markersize=30)

    plt.axhline(y=PEAKBW_KNL / CORE_KNL, color=COLOR_KNL, ls='--')
    plt.plot(bw_app, bw_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)

    plt.axhline(y=PEAKBW_KNM / CORE_KNM, color=COLOR_KNM, ls='--')
    plt.plot(bw_app, bw_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

    figure_save("bandwidth_per_core.png")


# Bandwidth per core used
figure_setup("Memory Throughput (per used core)", "(GB/sec)", (10,10), 0.25)

bw_app = []
bw_xeon = []
bw_knl = []
bw_knm = []
for app, rows in result_dic.items():
    xeon_core_used = get_core_used(rows[TITLE_XEON])
    knl_core_used = get_core_used(rows[TITLE_KNL])
    knm_core_used = get_core_used(rows[TITLE_KNM])

    bw_app.append(app[0:5])
    bw_xeon.append((get_bandwidth(rows[TITLE_XEON]) / xeon_core_used) if (xeon_core_used > 0) else BOTTOM)
    bw_knl.append((get_bandwidth(rows[TITLE_KNL])  / knl_core_used) if (knl_core_used > 0) else BOTTOM)
    bw_knm.append((get_bandwidth(rows[TITLE_KNM])  / knm_core_used) if (knm_core_used > 0) else BOTTOM)

plt.xticks(list(range(len(bw_app))), bw_app)
bw_app=list(range(len(bw_app)))

plt.axhline(y=PEAKBW_XEON / CORE_XEON, color=COLOR_XEON, ls='--')
plt.plot(bw_app, bw_xeon, ".", color=COLOR_XEON, label="Xeon", markersize=30)

plt.axhline(y=PEAKBW_KNL / CORE_KNL, color=COLOR_KNL, ls='--')
plt.plot(bw_app, bw_knl, ".", color=COLOR_KNL, label="KNL", markersize=30)

plt.axhline(y=PEAKBW_KNM / CORE_KNM, color=COLOR_KNM, ls='--')
plt.plot(bw_app, bw_knm, ".", color=COLOR_KNM, label="KNM", markersize=30)

figure_save("bandwidth_per_core_used.png")


# Memory/Backend bound
title="Bounded by Memory(Xeon=blue) / Back-end(KNL=green,KNM=red)"
ylabel="Percentage [%]"
figure_setup(title, ylabel, (12, 9), 10)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
bound1 = []
bound2 = []
bound3 = []
for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append("")
    ratio_label.append(app[:10])
    ratio_label.append("")
    ratio_label.append("")

    bound1.append(BOTTOM)
    bound1.append(get_boundedness(rows[TITLE_XEON]))
    bound1.append(BOTTOM)
    bound1.append(BOTTOM)
    bound1.append(BOTTOM)

    bound2.append(BOTTOM)
    bound2.append(BOTTOM)
    bound2.append(get_boundedness(rows[TITLE_KNL]))
    bound2.append(BOTTOM)
    bound2.append(BOTTOM)

    bound3.append(BOTTOM)
    bound3.append(BOTTOM)
    bound3.append(BOTTOM)
    bound3.append(get_boundedness(rows[TITLE_KNM]))
    bound3.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=60)
plt.xlim(xmin=0,xmax=split)
plt.ylim(ymax=100)
plt.bar(ratio_app[:split], bound1[:split], 0.8, label="INT")
plt.bar(ratio_app[:split], bound2[:split], 0.8, label="FP32")
plt.bar(ratio_app[:split], bound3[:split], 0.8, label="FP64")

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=60)
plt.xlim(xmin=split,xmax=2*split)
plt.ylim(ymax=100)
plt.bar(ratio_app[split:], bound1[split:], 0.8, label="INT")
plt.bar(ratio_app[split:], bound2[split:], 0.8, label="FP32")
plt.bar(ratio_app[split:], bound3[split:], 0.8, label="FP64")

figure_save("mem-bound.png")



# Ratio
title="FLOP to IOP Ratio in Solver [INT=blue, FP(SP)=green, FP(DP)=red]"
ylabel="Percentage [%]"
figure_setup(title, ylabel, (15, 9), 10)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
ratio_fp64 = []
ratio_fp32 = []
ratio_int = []
for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append(app[0:7] + "(BDW)")
    ratio_label.append(app[0:7] + "(KNL)")
    ratio_label.append(app[0:7] + "(KNM)")
    ratio_label.append("")

    ratio_xeon = get_ratio(rows[TITLE_XEON])
    ratio_knl  = get_ratio(rows[TITLE_KNL])
    ratio_knm  = get_ratio(rows[TITLE_KNM])

    ratio_fp64.append(BOTTOM)
    ratio_fp64.append(ratio_xeon[0])
    ratio_fp64.append(ratio_knl[0])
    ratio_fp64.append(ratio_knm[0])
    ratio_fp64.append(BOTTOM)

    ratio_fp32.append(BOTTOM)
    ratio_fp32.append(ratio_xeon[0] + ratio_xeon[1])
    ratio_fp32.append(ratio_knl[0] + ratio_knl[1])
    ratio_fp32.append(ratio_knm[0] + ratio_knm[1])
    ratio_fp32.append(BOTTOM)

    ratio_int.append(BOTTOM)
    ratio_int.append(100)
    ratio_int.append(100)
    ratio_int.append(100)
    ratio_int.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=16, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=60)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], ratio_int[:split], 0.8, label="INT")
plt.bar(ratio_app[:split], ratio_fp32[:split], 0.8, label="FP32")
plt.bar(ratio_app[:split], ratio_fp64[:split], 0.8, label="FP64")

plt.subplot(212)
plt.title(title, fontsize=16, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=60)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], ratio_int[split:], 0.8, label="INT")
plt.bar(ratio_app[split:], ratio_fp32[split:], 0.8, label="FP32")
plt.bar(ratio_app[split:], ratio_fp64[split:], 0.8, label="FP64")

figure_save("fp_to_int_ops_ratio.png")



# Hit rate
title="LLC Hit rate"
ylabel="(%)"
figure_setup(title, ylabel, (12, 9), 10)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
ratio_hit = []
ratio_ref = []
for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    ratio_hit.append(BOTTOM)
    ratio_hit.append(get_hitrate(rows[TITLE_XEON]))
    ratio_hit.append(get_hitrate(rows[TITLE_KNL]))
    ratio_hit.append(get_hitrate(rows[TITLE_KNM]))
    ratio_hit.append(BOTTOM)

    ratio_ref.append(BOTTOM)
    ratio_ref.append(100)
    ratio_ref.append(100)
    ratio_ref.append(100)
    ratio_ref.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.ylim(ymax=100)
plt.bar(ratio_app[:split], ratio_ref[:split], 0.8, label="REF")
plt.bar(ratio_app[:split], ratio_hit[:split], 0.8, label="HIT")

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.ylim(ymax=100)
plt.bar(ratio_app[split:], ratio_ref[split:], 0.8, label="REF")
plt.bar(ratio_app[split:], ratio_hit[split:], 0.8, label="HIT")

figure_save("llc_hitrate.png")


# Runtime rate
title="Solver Runtime vs Application Runtime"
ylabel="(%)"
figure_setup(title, ylabel, (12, 9), 10)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
ratio_sol = []
ratio_all = []
for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    ratio_sol.append(BOTTOM)
    ratio_sol.append(100 * get_solver_time(rows[TITLE_XEON]) / get_total_time(rows[TITLE_XEON]))
    ratio_sol.append(100 * get_solver_time(rows[TITLE_KNL]) / get_total_time(rows[TITLE_KNL]))
    ratio_sol.append(100 * get_solver_time(rows[TITLE_KNM]) / get_total_time(rows[TITLE_KNM]))
    ratio_sol.append(BOTTOM)

    ratio_all.append(BOTTOM)
    ratio_all.append(100)
    ratio_all.append(100)
    ratio_all.append(100)
    ratio_all.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], ratio_all[:split], 0.8, label="ENTIRE")
plt.bar(ratio_app[:split], ratio_sol[:split], 0.8, label="SOLVER")

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], ratio_all[split:], 0.8, label="ENTIRE")
plt.bar(ratio_app[split:], ratio_sol[split:], 0.8, label="SOLVER")

figure_save("solver_runtime.png")



# Energy
title="Power Consumption (package)"
ylabel="Watt"
figure_setup(title, ylabel, (12, 9), 10000)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
energy1 = []
energy2 = []
energy3 = []
acc = 1.0
for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    if (get_energy(rows[TITLE_KNM]) > 0):
        acc = acc * get_energy(rows[TITLE_KNL]) / get_energy(rows[TITLE_KNM])

    energy1.append(BOTTOM)
    energy1.append(get_energy(rows[TITLE_XEON]))
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)

    energy2.append(BOTTOM)
    energy2.append(BOTTOM)
    energy2.append(get_energy(rows[TITLE_KNL]))
    energy2.append(BOTTOM)
    energy2.append(BOTTOM)

    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(get_energy(rows[TITLE_KNM]))
    energy3.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.ylim(ymax=275)
plt.bar(ratio_app[:split], energy1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], energy2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], energy3[:split], 0.8, label="KNM", color=COLOR_KNM)

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.ylim(ymax=275)
plt.bar(ratio_app[split:], energy1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], energy2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], energy3[split:], 0.8, label="KNM", color=COLOR_KNM)

figure_save("energy.png")



# Energy per core used
title="Power Consumption (per used core)"
ylabel="Watt"
figure_setup(title, ylabel, (12, 9), 100)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
energy1 = []
energy2 = []
energy3 = []
acc = 1.0
for app, rows in result_dic.items():
    xeon_core_used = get_core_used(rows[TITLE_XEON])
    knl_core_used = get_core_used(rows[TITLE_KNL])
    knm_core_used = get_core_used(rows[TITLE_KNM])

    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    f1 = (get_energy(rows[TITLE_KNL]) / knl_core_used)
    f2 = (get_energy(rows[TITLE_KNM]) / knm_core_used)

    if (get_energy(rows[TITLE_KNM]) > 0):
        acc = acc * f1 / f2

    energy1.append(BOTTOM)
    energy1.append(get_energy(rows[TITLE_XEON]) / xeon_core_used)
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)

    energy2.append(BOTTOM)
    energy2.append(BOTTOM)
    energy2.append(get_energy(rows[TITLE_KNL]) / knl_core_used)
    energy2.append(BOTTOM)
    energy2.append(BOTTOM)

    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(get_energy(rows[TITLE_KNM]) / knm_core_used)
    energy3.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], energy1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], energy2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], energy3[:split], 0.8, label="KNM", color=COLOR_KNM)

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], energy1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], energy2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], energy3[split:], 0.8, label="KNM", color=COLOR_KNM)

figure_save("energy_per_used_core.png")


# Energy per time to solution
title="Power Consumption (per time-to-solution)"
ylabel="Joule"
figure_setup(title, ylabel, (12, 9), 100)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
energy1 = []
energy2 = []
energy3 = []
acc = 1.0
for app, rows in result_dic.items():
    xeon_solver_time = get_solver_time(rows[TITLE_XEON])
    knl_solver_time = get_solver_time(rows[TITLE_KNL])
    knm_solver_time = get_solver_time(rows[TITLE_KNM])

    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    f1 = (get_energy(rows[TITLE_KNL]) / knl_solver_time)
    f2 = (get_energy(rows[TITLE_KNM]) / knm_solver_time)

    if (get_energy(rows[TITLE_KNM]) > 0):
        acc = acc * f1 / f2

    energy1.append(BOTTOM)
    energy1.append(get_energy(rows[TITLE_XEON]) * xeon_solver_time)
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)
    energy1.append(BOTTOM)

    energy2.append(BOTTOM)
    energy2.append(BOTTOM)
    energy2.append(get_energy(rows[TITLE_KNL]) * knl_solver_time)
    energy2.append(BOTTOM)
    energy2.append(BOTTOM)

    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(BOTTOM)
    energy3.append(get_energy(rows[TITLE_KNM]) * knm_solver_time)
    energy3.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=80000)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], energy1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], energy2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], energy3[:split], 0.8, label="KNM", color=COLOR_KNM)

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=80000)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], energy1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], energy2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], energy3[split:], 0.8, label="KNM", color=COLOR_KNM)

figure_save("energy_per_tts.png")

figure_setup(title, ylabel, (12, 9), 100)
plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=11000)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], energy1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], energy2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], energy3[:split], 0.8, label="KNM", color=COLOR_KNM)
plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=11000)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], energy1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], energy2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], energy3[split:], 0.8, label="KNM", color=COLOR_KNM)
figure_save("energy_per_tts_zoom.png")


# Byte/Flop
title="Bytes/Flop Ratio"
ylabel="Byte/Flop"
figure_setup(title, ylabel, (12, 9), 0.25)

ratio_app = list(range(len(result_dic) * 5))
split = int(len(result_dic) * 5 / 2)
ratio_label = []
bf1 = []
bf2 = []
bf3 = []

for app, rows in result_dic.items():
    ratio_label.append("")
    ratio_label.append(app[0:5] + "(XEON)")
    ratio_label.append(app[0:5] + "(KNL)")
    ratio_label.append(app[0:5] + "(KNM)")
    ratio_label.append("")

    bf1.append(BOTTOM)
    bf1.append(get_bf(rows[TITLE_XEON]))
    bf1.append(BOTTOM)
    bf1.append(BOTTOM)
    bf1.append(BOTTOM)

    bf2.append(BOTTOM)
    bf2.append(BOTTOM)
    bf2.append(get_bf(rows[TITLE_KNL]))
    bf2.append(BOTTOM)
    bf2.append(BOTTOM)

    bf3.append(BOTTOM)
    bf3.append(BOTTOM)
    bf3.append(BOTTOM)
    bf3.append(get_bf(rows[TITLE_KNM]))
    bf3.append(BOTTOM)

plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=450000000)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], bf1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], bf2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], bf3[:split], 0.8, label="KNM", color=COLOR_KNM)

plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=450000000)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], bf1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], bf2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], bf3[split:], 0.8, label="KNM", color=COLOR_KNM)

figure_save("byte_per_flop.png")

figure_setup(title, ylabel, (12, 9), 0.25)
plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=120)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], bf1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], bf2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], bf3[:split], 0.8, label="KNM", color=COLOR_KNM)
plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=120)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], bf1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], bf2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], bf3[split:], 0.8, label="KNM", color=COLOR_KNM)
figure_save("byte_per_flop_zoom1.png")

figure_setup(title, ylabel, (12, 9), 0.25)
plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=20)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], bf1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], bf2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], bf3[:split], 0.8, label="KNM", color=COLOR_KNM)
plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=20)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], bf1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], bf2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], bf3[split:], 0.8, label="KNM", color=COLOR_KNM)
figure_save("byte_per_flop_zoom2.png")

figure_setup(title, ylabel, (12, 9), 0.25)
plt.subplot(211)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=2)
plt.xticks(ratio_app[:split], ratio_label[:split], rotation=90)
plt.xlim(xmin=0,xmax=split)
plt.bar(ratio_app[:split], bf1[:split], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[:split], bf2[:split], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[:split], bf3[:split], 0.8, label="KNM", color=COLOR_KNM)
plt.subplot(212)
plt.title(title, fontsize=20, fontweight='bold')
plt.ylabel(ylabel)
plt.ylim(ymax=2)
plt.xticks(ratio_app[split:], ratio_label[split:], rotation=90)
plt.xlim(xmin=split,xmax=2*split)
plt.bar(ratio_app[split:], bf1[split:], 0.8, label="XEON", color=COLOR_XEON)
plt.bar(ratio_app[split:], bf2[split:], 0.8, label="KNL", color=COLOR_KNL)
plt.bar(ratio_app[split:], bf3[split:], 0.8, label="KNM", color=COLOR_KNM)
figure_save("byte_per_flop_zoom3.png")


# Clean
os.unlink(XLSX_FILE)

